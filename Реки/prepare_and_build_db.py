# -*- coding: utf-8 -*-
"""
Полный конвейер:
  ШАГ 1 - Раскидать CSV-файлы гидропостов по папкам-рекам (по данным из Excel)
  ШАГ 2 - Создать единую SQLite-базу с гидропостами и метеостанциями

Источники:
  Гидропосты/          - CSV с температурами (gidro_num -> река из Excel)
  Свод - гидропосты.xlsx - маппинг gidro_num -> namewater (река)
  Метеостанции_1/      - CSV с погодой (часть 1)
  Метеостанции_2/      - CSV с погодой (часть 2)
  Список метеостанций.xlsx - справочник метеостанций
"""

import psycopg2
import csv
import os
import sys
import re
import time
import shutil
import openpyxl
from pathlib import Path


# ──────────────────────────────────────────────
#  Конфигурация
# ──────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'hydro_meteo.db'
def get_db_connection():
    import os
    db_url = os.environ.get('DATASET_DATABASE_URL', os.environ.get('DATABASE_URL', 'postgresql://portal_user:admin135@localhost:5432/portal_db'))
    conn = psycopg2.connect(db_url)
    conn.autocommit = True
    with conn.cursor() as cur:
        cur.execute("CREATE SCHEMA IF NOT EXISTS dataset_schema"
        cur.execute("SET search_path TO dataset_schema")
    # For executing large inserts we want to control transaction if possible, or rely on autocommit. 
    # But executemany without transaction might be slow.
    # We will just disable autocommit in the caller.
    return conn

HYDRO_SRC_DIR = BASE_DIR / "Гидропосты"
HYDRO_EXCEL = BASE_DIR / "Свод - гидропосты.xlsx"
HYDRO_SORTED_DIR = BASE_DIR / "Гидропосты_по_рекам"

METEO_SRC_DIRS = [BASE_DIR / "Метеостанции_1", BASE_DIR / "Метеостанции_2"]
METEO_EXCEL = BASE_DIR / "Список метеостанций.xlsx"
METEO_SORTED_DIR = BASE_DIR / "Метеостанции_по_типам"

CSV_ENCODING = "utf-8"
BATCH_SIZE = 5000


# Фикс вывода для консоли Windows
if sys.stdout.encoding and sys.stdout.encoding.lower( != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass


# ═══════════════════════════════════════════════
#  ШАГ 1: Раскидать гидропосты по папкам-рекам
# ═══════════════════════════════════════════════

def sanitize_folder_name(name):
    """Очищает имя для использования как папка."""
    # Убираем/заменяем символы, недопустимые в путях Windows
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', ' ', name)
    return name


def load_hydro_mapping():
    """Читает Excel и строит маппинг gidro_num -> namewater (река)."""
    wb = openpyxl.load_workbook(str(HYDRO_EXCEL))
    ws = wb.worksheets[0]

    mapping = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        gidro_num = row[0]   # gidro_num
        namewater = row[2]   # namewater (река)
        if gidro_num and namewater:
            # Нормализуем ключ: "3027.0" -> "3027"
            key = str(gidro_num).replace('.0', '')
            if key not in mapping:
                mapping[key] = str(namewater).strip()

    wb.close()
    return mapping


def sort_hydro_files():
    """Раскидывает CSV-файлы из Гидропосты/ по папкам-рекам."""
    print("=" * 60)
    print("  ШАГ 1: Сортировка гидропостов по рекам")
    print("=" * 60)

    if not HYDRO_SRC_DIR.exists():
        print("[!] Папка Гидропосты не найдена: %s" % HYDRO_SRC_DIR)
        return False

    if not HYDRO_EXCEL.exists():
        print("[!] Excel с маппингом не найден: %s" % HYDRO_EXCEL)
        return False

    # Загружаем маппинг
    mapping = load_hydro_mapping()
    print("  Маппинг загружен: %d записей" % len(mapping))

    # Создаём целевую папку
    HYDRO_SORTED_DIR.mkdir(exist_ok=True)

    # Собираем CSV файлы (также файлы без расширения, как 3812.0_1)
    all_files = list(HYDRO_SRC_DIR.iterdir())
    csv_files = [f for f in all_files if f.is_file() and f.suffix in ('.csv', '')]

    # Исключаем Excel файлы
    csv_files = [f for f in csv_files if not f.name.endswith('.xlsx')]

    sorted_count = 0
    skipped = []
    river_counts = {}

    for csv_file in sorted(csv_files):
        # Извлекаем gidro_num из имени файла: "3027.0_1.csv" -> "3027"
        fname = csv_file.stem if csv_file.suffix == '.csv' else csv_file.name
        match = re.match(r'(\d+\.?\d*)', fname)
        if not match:
            skipped.append(csv_file.name)
            continue

        gnum = match.group(1).replace('.0', '')

        if gnum not in mapping:
            skipped.append(csv_file.name)
            continue

        river_name = mapping[gnum]
        folder_name = sanitize_folder_name(river_name)
        target_dir = HYDRO_SORTED_DIR / folder_name
        target_dir.mkdir(exist_ok=True)

        # Копируем файл
        target_path = target_dir / csv_file.name
        if not target_path.exists():
            shutil.copy2(str(csv_file), str(target_path))

        sorted_count += 1
        river_counts[river_name] = river_counts.get(river_name, 0) + 1

    print("")
    print("  Результат:")
    print("    Файлов обработано: %d" % sorted_count)
    print("    Пропущено:        %d" % len(skipped))
    print("    Рек (папок):      %d" % len(river_counts))

    if skipped:
        print("    Пропущенные файлы: %s" % ', '.join(skipped[:10]))

    print("")
    print("  Рек по количеству файлов:")
    for river, count in sorted(river_counts.items(), key=lambda x: -x[1])[:20]:
        print("    %s: %d файлов" % (river, count))
    if len(river_counts) > 20:
        print("    ... и ещё %d рек" % (len(river_counts) - 20))

    print("")
    print("  [OK] Файлы разложены в: %s" % HYDRO_SORTED_DIR)
    return True


# ═══════════════════════════════════════════════
#  ШАГ 1б: Раскидать метеостанции по папкам-типам
# ═══════════════════════════════════════════════

METEO_CATEGORY_NAMES = {
    'station_meteo': 'Метеорологические',
    'station_gidro': 'Гидрологические',
    'station_aero': 'Аэрологические',
    'station_mor': 'Морские',
    'station_agro': 'Агрометеорологические',
    'post_meteo': 'Метеопосты',
}


def load_meteo_category_mapping():
    """Читает Excel (лист 2) и строит маппинг station_id -> ktocategory."""
    wb = openpyxl.load_workbook(str(METEO_EXCEL))
    # Используем лист 2 (с полными данными включая ktocategory)
    ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    mapping = {}  # station_id -> ktocategory
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        d = {str(header[i]): row[i] for i in range(min(len(header), len(row)))}
        sid = d.get('stationid')
        cat = d.get('ktocategory')
        if sid:
            mapping[int(sid)] = str(cat) if cat else 'unknown'

    wb.close()
    return mapping


def sort_meteo_files():
    """Раскидывает CSV-файлы метеостанций по папкам-типам."""
    print("")
    print("=" * 60)
    print("  ШАГ 1б: Сортировка метеостанций по типам")
    print("=" * 60)

    if not METEO_EXCEL.exists():
        print("  [!] Excel метеостанций не найден: %s" % METEO_EXCEL)
        return False

    # Загружаем маппинг station_id -> ktocategory
    cat_mapping = load_meteo_category_mapping()
    print("  Маппинг загружен: %d записей" % len(cat_mapping))

    # Создаём целевую папку
    METEO_SORTED_DIR.mkdir(exist_ok=True)

    sorted_count = 0
    skipped = []
    category_counts = {}

    for meteo_dir in METEO_SRC_DIRS:
        if not meteo_dir.exists():
            print("  [!] Папка не найдена: %s" % meteo_dir)
            continue

        csv_files = [f for f in meteo_dir.iterdir() if f.is_file() and f.suffix == '.csv']

        for csv_file in sorted(csv_files):
            # Извлекаем station_id из имени файла: "21432_1.csv" -> 21432
            match = re.match(r'(\d+)', csv_file.name)
            if not match:
                skipped.append(csv_file.name)
                continue

            station_id = int(match.group(1))

            if station_id not in cat_mapping:
                # Не найден в Excel — кладём в 'Прочие'
                category = 'unknown'
            else:
                category = cat_mapping[station_id]

            # Человеко-читаемое имя папки
            folder_name = METEO_CATEGORY_NAMES.get(category, category)
            target_dir = METEO_SORTED_DIR / folder_name
            target_dir.mkdir(exist_ok=True)

            target_path = target_dir / csv_file.name
            if not target_path.exists():
                shutil.copy2(str(csv_file), str(target_path))

            sorted_count += 1
            category_counts[folder_name] = category_counts.get(folder_name, 0) + 1

    print("")
    print("  Результат:")
    print("    Файлов обработано: %d" % sorted_count)
    print("    Пропущено:        %d" % len(skipped))
    print("    Типов (папок):    %d" % len(category_counts))

    if skipped:
        print("    Пропущенные файлы: %s" % ', '.join(skipped[:10]))

    print("")
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        print("    %s: %d файлов" % (cat, count))

    print("")
    print("  [OK] Файлы разложены в: %s" % METEO_SORTED_DIR)
    return True


# ═══════════════════════════════════════════════
#  ШАГ 2: Создание базы данных
# ═══════════════════════════════════════════════

def create_database(conn):
    """Создаёт все таблицы в базе данных."""
    cursor = conn.cursor()

    # --- Гидропосты ---

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS rivers (
            id          INTEGER PRIMARY KEY ,
            name        TEXT NOT NULL UNIQUE
        
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hydro_stations (
            id          INTEGER PRIMARY KEY ,
            gidro_num   INTEGER NOT NULL,
            identifier  INTEGER NOT NULL,
            river_id    INTEGER NOT NULL,
            FOREIGN KEY (river_id REFERENCES rivers(id),
            UNIQUE(gidro_num, identifier)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hydro_temperatures (
            id          INTEGER PRIMARY KEY ,
            station_id  INTEGER NOT NULL,
            dt          TEXT NOT NULL,
            temp        REAL,
            temptype    TEXT NOT NULL,
            source_file TEXT,
            FOREIGN KEY (station_id REFERENCES hydro_stations(id)
        )
    """)

    # --- Метеостанции ---

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS meteo_stations (
            id                  INTEGER PRIMARY KEY ,
            station_identifier  INTEGER NOT NULL UNIQUE,
            station_name        TEXT,
            name_long           TEXT
        
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS meteo_observations (
            id                              INTEGER PRIMARY KEY ,
            station_id                      INTEGER NOT NULL,
            observation_date                TEXT,
            observation_ts                  TEXT,
            observation_period              INTEGER,
            present_weather                 REAL,
            wind_direction                  REAL,
            wind_speed                      REAL,
            max_wind_gust_speed             REAL,
            air_temperature                 REAL,
            dewpoint_temperature            REAL,
            ground_min_temp_past_12h        REAL,
            source_file                     TEXT,
            FOREIGN KEY (station_id REFERENCES meteo_stations(id)
        )
    """)

    # --- Индексы ---
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_hydro_station ON hydro_temperatures(station_id")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_hydro_dt ON hydro_temperatures(dt")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_hydro_type ON hydro_temperatures(temptype")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_hydro_gidro ON hydro_stations(gidro_num")

    cursor.execute("CREATE INDEX IF NOT EXISTS idx_meteo_station ON meteo_observations(station_id")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_meteo_date ON meteo_observations(observation_date")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_meteo_ts ON meteo_observations(observation_ts")

    conn.commit()
    print("  [OK] Таблицы и индексы созданы")


def load_meteo_station_mapping():
    """Загружает справочник метеостанций из Excel."""
    if not METEO_EXCEL.exists():
        print("  [!] Excel метеостанций не найден: %s" % METEO_EXCEL)
        return {}

    wb = openpyxl.load_workbook(str(METEO_EXCEL))
    ws = wb.worksheets[0]

    mapping = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        station_id = row[1]   # stationid
        name_long = row[2]    # namelong
        if station_id:
            mapping[int(station_id)] = str(name_long) if name_long else ''

    wb.close()
    return mapping


# --- Гидропосты: импорт ---

def get_or_create_river(cursor, river_name):
    cursor.execute("SELECT id FROM rivers WHERE name = %s", (river_name,)
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute("INSERT INTO rivers (name VALUES (?)", (river_name,))
    return cursor.lastrowid


def get_or_create_hydro_station(cursor, gidro_num, identifier, river_id):
    cursor.execute(
        "SELECT id FROM hydro_stations WHERE gidro_num = %s AND identifier = %s",
        (gidro_num, identifier
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute(
        "INSERT INTO hydro_stations (gidro_num, identifier, river_id VALUES (?, ?, ?)",
        (gidro_num, identifier, river_id)
    )
    return cursor.lastrowid


def import_hydro_csv(cursor, filepath, river_id, station_cache):
    """Импортирует один CSV-файл гидропоста. Возвращает кол-во строк."""
    inserted = 0
    filename = filepath.name
    batch = []

    with open(filepath, "r", encoding=CSV_ENCODING) as f:
        reader = csv.DictReader(f)
        for row in reader:
            gidro_num = int(float(row["gidro_num"]))
            identifier = int(float(row["identifier"]))
            dt = row["dt"].strip()
            temptype = row["temptype"].strip()

            try:
                temp = float(row["temp"])
            except (ValueError, TypeError):
                temp = None

            station_key = (gidro_num, identifier)
            if station_key not in station_cache:
                station_id = get_or_create_hydro_station(cursor, gidro_num, identifier, river_id)
                station_cache[station_key] = station_id
            else:
                station_id = station_cache[station_key]

            batch.append((station_id, dt, temp, temptype, filename))
            inserted += 1

            if len(batch) >= BATCH_SIZE:
                cursor.executemany(
                    "INSERT INTO hydro_temperatures (station_id, dt, temp, temptype, source_file VALUES (?, ?, ?, ?, ?)",
                    batch
                )
                batch.clear()

    if batch:
        cursor.executemany(
            "INSERT INTO hydro_temperatures (station_id, dt, temp, temptype, source_file VALUES (?, ?, ?, ?, ?)",
            batch
        )

    return inserted


def import_all_hydro(conn):
    """Импортирует все гидропосты из папок-рек."""
    print("")
    print("-" * 60)
    print("  Импорт гидропостов")
    print("-" * 60)

    cursor = conn.cursor()
    station_cache = {}

    # Используем отсортированные папки если есть, иначе — исходную папку Гидропосты
    if HYDRO_SORTED_DIR.exists():
        source_dir = HYDRO_SORTED_DIR
        use_folder_as_river = True
    else:
        source_dir = HYDRO_SRC_DIR
        use_folder_as_river = False

    if not source_dir.exists():
        print("  [!] Нет данных для гидропостов")
        return 0

    total_files = 0
    total_rows = 0

    if use_folder_as_river:
        # Каждая подпапка = река
        river_dirs = sorted([d for d in source_dir.iterdir() if d.is_dir()])
        print("  Рек (папок): %d" % len(river_dirs))

        for river_dir in river_dirs:
            river_name = river_dir.name
            river_id = get_or_create_river(cursor, river_name)
            csv_files = sorted(list(river_dir.glob("*.csv")) + [f for f in river_dir.iterdir() if f.is_file() and f.suffix == ''])

            if not csv_files:
                continue

            river_rows = 0
            for csv_file in csv_files:
                try:
                    rows = import_hydro_csv(cursor, csv_file, river_id, station_cache)
                    river_rows += rows
                    total_files += 1
                    sys.stdout.write(".")
                    sys.stdout.flush()
                except Exception as e:
                    print("\n  [!] Ошибка %s: %s" % (csv_file.name, e))

            conn.commit()
            total_rows += river_rows
        print("")
    else:
        # Все файлы в одной папке — используем маппинг
        mapping = load_hydro_mapping()
        csv_files = sorted([f for f in source_dir.iterdir() if f.is_file() and not f.name.endswith('.xlsx')])

        for csv_file in csv_files:
            fname = csv_file.stem if csv_file.suffix == '.csv' else csv_file.name
            match = re.match(r'(\d+\.?\d*)', fname)
            if not match:
                continue
            gnum = match.group(1).replace('.0', '')
            river_name = mapping.get(gnum, 'НЕИЗВЕСТНАЯ')
            river_id = get_or_create_river(cursor, river_name)

            try:
                rows = import_hydro_csv(cursor, csv_file, river_id, station_cache)
                total_rows += rows
                total_files += 1
                sys.stdout.write(".")
                sys.stdout.flush()
            except Exception as e:
                print("\n  [!] Ошибка %s: %s" % (csv_file.name, e))

        conn.commit()
        print("")

    print("  [OK] Гидропосты: %s записей из %d файлов, %d станций" % (
        "{:,}".format(total_rows), total_files, len(station_cache)))
    return total_rows


# --- Метеостанции: импорт ---

def get_or_create_meteo_station(cursor, station_identifier, station_name, name_long, cache):
    if station_identifier in cache:
        return cache[station_identifier]

    cursor.execute(
        "SELECT id FROM meteo_stations WHERE station_identifier = %s",
        (station_identifier,
    )
    row = cursor.fetchone()
    if row:
        cache[station_identifier] = row[0]
        return row[0]

    cursor.execute(
        "INSERT INTO meteo_stations (station_identifier, station_name, name_long VALUES (?, ?, ?)",
        (station_identifier, station_name, name_long)
    )
    sid = cursor.lastrowid
    cache[station_identifier] = sid
    return sid


def safe_float(val):
    """Безопасный парсинг float."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """Безопасный парсинг int."""
    if val is None or val == '':
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def import_meteo_csv(cursor, filepath, meteo_name_map, station_cache):
    """Импортирует один CSV-файл метеостанции. Возвращает кол-во строк."""
    inserted = 0
    filename = filepath.name
    batch = []

    with open(filepath, "r", encoding=CSV_ENCODING) as f:
        reader = csv.DictReader(f)
        for row in reader:
            station_identifier = safe_int(row.get("station_identifier"))
            if station_identifier is None:
                continue

            station_name = row.get("station_name", "")
            name_long = meteo_name_map.get(station_identifier, station_name)

            station_id = get_or_create_meteo_station(
                cursor, station_identifier, station_name, name_long, station_cache
            )

            batch.append((
                station_id,
                row.get("observation_date", ""),
                row.get("observation_ts", ""),
                safe_int(row.get("observation_period")),
                safe_float(row.get("presentweather")),
                safe_float(row.get("winddirection")),
                safe_float(row.get("windspeed")),
                safe_float(row.get("maximumwindgustspeed")),
                safe_float(row.get("airtemperature")),
                safe_float(row.get("dewpointtemperature")),
                safe_float(row.get("groundminimumtemperaturepast12hours")),
                filename,
            ))
            inserted += 1

            if len(batch) >= BATCH_SIZE:
                cursor.executemany("""
                    INSERT INTO meteo_observations
                    (station_id, observation_date, observation_ts, observation_period,
                     present_weather, wind_direction, wind_speed, max_wind_gust_speed,
                     air_temperature, dewpoint_temperature, ground_min_temp_past_12h,
                     source_file
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, batch)
                batch.clear()

    if batch:
        cursor.executemany("""
            INSERT INTO meteo_observations
            (station_id, observation_date, observation_ts, observation_period,
             present_weather, wind_direction, wind_speed, max_wind_gust_speed,
             air_temperature, dewpoint_temperature, ground_min_temp_past_12h,
             source_file
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, batch)

    return inserted


def import_all_meteo(conn):
    """Импортирует все метеостанции из отсортированных или исходных папок."""
    print("")
    print("-" * 60)
    print("  Импорт метеостанций")
    print("-" * 60)

    cursor = conn.cursor()
    station_cache = {}

    # Загружаем справочник имён
    meteo_name_map = load_meteo_station_mapping()
    print("  Справочник метеостанций: %d записей" % len(meteo_name_map))

    total_files = 0
    total_rows = 0

    # Используем отсортированные папки если есть
    if METEO_SORTED_DIR.exists():
        type_dirs = sorted([d for d in METEO_SORTED_DIR.iterdir() if d.is_dir()])
        print("  Типов (папок): %d" % len(type_dirs))

        for type_dir in type_dirs:
            csv_files = sorted(type_dir.glob("*.csv"))
            if not csv_files:
                continue

            dir_rows = 0
            print("  %s: %d файлов" % (type_dir.name, len(csv_files)))

            for csv_file in csv_files:
                try:
                    rows = import_meteo_csv(cursor, csv_file, meteo_name_map, station_cache)
                    dir_rows += rows
                    total_files += 1
                    sys.stdout.write(".")
                    sys.stdout.flush()
                except Exception as e:
                    print("\n  [!] Ошибка %s: %s" % (csv_file.name, e))

            conn.commit()
            total_rows += dir_rows
            print("")
            print("  [OK] %s: %s записей" % (type_dir.name, "{:,}".format(dir_rows)))
    else:
        # Фоллбэк: используем исходные папки
        for meteo_dir in METEO_SRC_DIRS:
            if not meteo_dir.exists():
                print("  [!] Папка не найдена: %s" % meteo_dir)
                continue

            csv_files = sorted(meteo_dir.glob("*.csv"))
            print("  %s: %d файлов" % (meteo_dir.name, len(csv_files)))

            dir_rows = 0
            for csv_file in csv_files:
                try:
                    rows = import_meteo_csv(cursor, csv_file, meteo_name_map, station_cache)
                    dir_rows += rows
                    total_files += 1
                    sys.stdout.write(".")
                    sys.stdout.flush()
                except Exception as e:
                    print("\n  [!] Ошибка %s: %s" % (csv_file.name, e))

            conn.commit()
            total_rows += dir_rows
            print("")
            print("  [OK] %s: %s записей" % (meteo_dir.name, "{:,}".format(dir_rows)))

    print("  [OK] Метеостанции итого: %s записей из %d файлов, %d станций" % (
        "{:,}".format(total_rows), total_files, len(station_cache)))
    return total_rows


# ═══════════════════════════════════════════════
#  Сводка
# ═══════════════════════════════════════════════

def print_summary(conn):
    """Выводит сводку по базе данных."""
    cursor = conn.cursor()

    print("")
    print("=" * 60)
    print("  СВОДКА ПО БАЗЕ ДАННЫХ")
    print("=" * 60)

    # Реки
    cursor.execute("SELECT COUNT(* FROM rivers")
    river_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(* FROM hydro_stations")
    hstation_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(* FROM hydro_temperatures")
    htemp_count = cursor.fetchone()[0]

    print("")
    print("  ГИДРОПОСТЫ:")
    print("    Рек:       %d" % river_count)
    print("    Станций:   %d" % hstation_count)
    print("    Записей:   %s" % "{:,}".format(htemp_count))

    if htemp_count > 0:
        cursor.execute("SELECT MIN(dt, MAX(dt) FROM hydro_temperatures")
        dmin, dmax = cursor.fetchone()
        print("    Даты:      %s -- %s" % (dmin, dmax))

        cursor.execute("SELECT temptype, COUNT(* FROM hydro_temperatures GROUP BY temptype")
        for ttype, cnt in cursor.fetchall():
            print("    %s: %s" % (ttype, "{:,}".format(cnt)))

    # Топ-5 рек
    cursor.execute("""
        SELECT r.name, COUNT(t.id as cnt
        FROM hydro_temperatures t
        JOIN hydro_stations s ON t.station_id = s.id
        JOIN rivers r ON s.river_id = r.id
        GROUP BY r.name ORDER BY cnt DESC LIMIT 10
    """)
    rows = cursor.fetchall()
    if rows:
        print("")
        print("  Топ-10 рек по количеству записей:")
        for rname, cnt in rows:
            print("    %s: %s" % (rname, "{:,}".format(cnt)))

    # Метеостанции
    cursor.execute("SELECT COUNT(* FROM meteo_stations")
    mstation_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(* FROM meteo_observations")
    mobs_count = cursor.fetchone()[0]

    print("")
    print("  МЕТЕОСТАНЦИИ:")
    print("    Станций:   %d" % mstation_count)
    print("    Записей:   %s" % "{:,}".format(mobs_count))

    if mobs_count > 0:
        cursor.execute("SELECT MIN(observation_date, MAX(observation_date) FROM meteo_observations")
        dmin, dmax = cursor.fetchone()
        print("    Даты:      %s -- %s" % (dmin, dmax))

    # Примеры
    if htemp_count > 0:
        print("")
        print("  Примеры данных (гидро):")
        cursor.execute("""
            SELECT s.gidro_num, r.name, t.dt, t.temp, t.temptype
            FROM hydro_temperatures t
            JOIN hydro_stations s ON t.station_id = s.id
            JOIN rivers r ON s.river_id = r.id
            ORDER BY t.dt DESC LIMIT 5
        """
        for row in cursor.fetchall():
            print("    Пост %d (%s) | %s | %s C | %s" % row)

    if mobs_count > 0:
        print("")
        print("  Примеры данных (метео):")
        cursor.execute("""
            SELECT s.station_identifier, s.name_long, o.observation_date,
                   o.air_temperature, o.wind_speed
            FROM meteo_observations o
            JOIN meteo_stations s ON o.station_id = s.id
            ORDER BY o.observation_date DESC LIMIT 5
        """
        for row in cursor.fetchall():
            print("    Ст. %d (%s) | %s | t=%s C | v=%s м/с" % row)

    # Размер БД
    db_size_mb = DB_PATH.stat().st_size / (1024 * 1024)
    print("")
    print("  Размер БД: %.1f МБ" % db_size_mb)


# ═══════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════

def main():
    start = time.time()

    print("")
    print("=" * 60)
    print("  Подготовка данных и создание базы")
    print("  hydro_meteo.db")
    print("=" * 60)
    print("  Рабочая папка: %s" % BASE_DIR)
    print("")

    # ШАГ 1: Раскидать гидропосты по папкам-рекам
    sort_hydro_files()

    # ШАГ 1б: Раскидать метеостанции по папкам-типам
    sort_meteo_files()

    # ШАГ 2: Создать и наполнить БД
    print("")
    print("=" * 60)
    print("  ШАГ 2: Создание базы данных")
    print("=" * 60)

    if DB_PATH.exists():
        DB_PATH.unlink()
        print("  [*] Старая БД удалена")

    conn = get_db_connection()
    conn.autocommit = False
    conn.execute("PRAGMA journal_mode = WAL"
    conn.execute("PRAGMA synchronous = NORMAL"
    conn.execute("PRAGMA cache_size = -64000"
    conn.execute("PRAGMA encoding = 'UTF-8'"

    try:
        create_database(conn)
        hydro_rows = import_all_hydro(conn)
        meteo_rows = import_all_meteo(conn)
        print_summary(conn)
    except Exception as e:
        print("")
        print("[!!!] Критическая ошибка: %s" % e)
        import traceback
        traceback.print_exc()
    finally:
        conn.close()

    elapsed = time.time() - start
    print("")
    print("=" * 60)
    print("  [OK] Готово за %.1f сек" % elapsed)
    print("  БД: %s" % DB_PATH)
    print("=" * 60)


if __name__ == "__main__":
    main()
